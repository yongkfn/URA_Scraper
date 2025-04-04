import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import os
import logging
import time
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("ura_scraper.log"),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

class URAScraper:
    def __init__(self):
        self.base_url = "https://www.ura.gov.sg/Corporate/Land-Sales/Current-URA-GLS-Sites"
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5"
        }
        self.session = requests.Session()
        self.output_file = "ura_land_sales.xlsx"
        
    def fetch_main_page(self):
        """Fetch the main GLS sites page"""
        try:
            response = self.session.get(self.base_url, headers=self.headers)
            response.raise_for_status()
            logger.info(f"Successfully fetched main page. Status code: {response.status_code}")
            
            # No need to follow the "#Confirmed-List" link as it's just an anchor on the same page
            # Just return the current page content
            return response.text
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching main page: {e}")
            return None
            
    def parse_main_page(self, html):
        """Parse the main page to extract GLS sites information"""
        if not html:
            return []
            
        soup = BeautifulSoup(html, 'html.parser')
        sites_data = []
        
        # Look for the Confirmed List section or any table
        tables = soup.select('table.table, table.gls-table, table')
        
        if not tables:
            logger.warning("No tables found on the page")
            return []
        
        # Try each table until we find one with relevant data
        for table in tables:
            # Check if this table has the expected structure
            headers = [th.get_text(strip=True).lower() for th in table.select('thead th, tr th')]
            
            if not headers and table.select('tr'):
                # If no explicit headers, try to use the first row as headers
                first_row = table.select('tr')[0]
                headers = [td.get_text(strip=True).lower() for td in first_row.select('td')]
            
            logger.info(f"Found table with headers: {headers}")
            
            # Check if headers match what we're looking for
            has_required_columns = False
            
            # Expected columns: No, Location, Site Area, Gross Plot Ratio, Status
            if headers and any('no' in h.lower() for h in headers) and any('location' in h.lower() for h in headers):
                has_required_columns = True
            
            if has_required_columns:
                logger.info("Found table with required columns")
                
                # Get all rows, skipping the header row
                all_rows = table.select('tr')
                rows = all_rows[1:] if len(all_rows) > 0 else []
                
                # If there's a tbody, use those rows instead
                tbody_rows = table.select('tbody tr')
                if tbody_rows:
                    rows = tbody_rows
                
                logger.info(f"Found {len(rows)} data rows")
                
                # Map column indices
                no_col = -1
                location_col = -1
                area_col = -1
                gpr_col = -1
                status_col = -1
                
                for i, header in enumerate(headers):
                    if 'no' == header.lower() or 'no.' == header.lower():
                        no_col = i
                    elif 'location' in header.lower():
                        location_col = i
                    elif 'site area' in header.lower() or 'area' in header.lower():
                        area_col = i
                    elif 'gross plot ratio' in header.lower() or 'plot ratio' in header.lower():
                        gpr_col = i
                    elif 'status' in header.lower():
                        status_col = i
                
                for row in rows:
                    cells = row.select('td')
                    
                    # Skip rows without enough columns
                    if len(cells) < 3:  # At minimum we need location and status
                        continue
                    
                    try:
                        # Extract site number
                        site_no = ""
                        if no_col >= 0 and no_col < len(cells):
                            site_no = cells[no_col].get_text(strip=True)
                        
                        # Extract location (site name)
                        location = ""
                        if location_col >= 0 and location_col < len(cells):
                            location = cells[location_col].get_text(strip=True)
                        
                        # If we don't have a location, this might not be a data row
                        if not location:
                            # Check if this is a category header row like "Residential Sites"
                            if len(cells) == 1 and cells[0].get('colspan'):
                                category_text = cells[0].get_text(strip=True)
                                logger.info(f"Found category header: {category_text}")
                                # No need to process this row further
                                continue
                            else:
                                # Skip rows without location
                                continue
                        
                        # Extract site area
                        site_area = ""
                        if area_col >= 0 and area_col < len(cells):
                            site_area = cells[area_col].get_text(strip=True)
                        
                        # Extract gross plot ratio
                        gpr = ""
                        if gpr_col >= 0 and gpr_col < len(cells):
                            gpr = cells[gpr_col].get_text(strip=True)
                        
                        # Extract status
                        status = ""
                        if status_col >= 0 and status_col < len(cells):
                            status = cells[status_col].get_text(strip=True)
                        
                        # Check if there's a link in the location cell
                        link = None
                        if location_col >= 0 and location_col < len(cells):
                            location_cell = cells[location_col]
                            if location_cell.find('a'):
                                href = location_cell.find('a').get('href')
                                if href:
                                    if href.startswith('/'):
                                        link = f"https://www.ura.gov.sg{href}"
                                    elif href.startswith('http'):
                                        link = href
                                    else:
                                        link = f"https://www.ura.gov.sg/{href}"
                        
                        # Add the site to our data in the exact format needed
                        sites_data.append({
                            'No': site_no,
                            'Location': location,
                            'Site Area (Ha)': site_area,
                            'Gross Plot Ratio': gpr,
                            'Status': status,
                            'Link': link,
                            'scraped_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        })
                        logger.info(f"Added site: {location}, Status: {status}")
                    except Exception as e:
                        logger.error(f"Error parsing row: {e}")
                        continue
                
                # If we found some sites in this table, no need to check others
                if sites_data:
                    logger.info(f"Successfully extracted {len(sites_data)} sites from table")
                    break
        
        logger.info(f"Found {len(sites_data)} sites on the page")
        return sites_data
        
    def fetch_project_details(self, sites_data):
        """Fetch details for awarded projects"""
        for site in sites_data:
            if site['Status'].lower() == 'awarded' and site['Link']:
                logger.info(f"Fetching details for awarded site: {site['Location']}")
                
                try:
                    # Add a delay to avoid overwhelming the server
                    time.sleep(2)
                    
                    response = self.session.get(site['Link'], headers=self.headers)
                    response.raise_for_status()
                    
                    # Store the fact that we followed the link for this site
                    site['details'] = {'followed_link': True}
                    logger.info(f"Successfully fetched details for {site['Location']}")
                    
                except Exception as e:
                    logger.error(f"Error fetching details for {site['Location']}: {e}")
                    site['details'] = {'error': str(e)}
        
        return sites_data
                
    def save_data(self, sites_data):
        """Save the scraped data to an Excel file with proper formatting"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "URA GLS Sites"
            
            # Define the columns exactly as requested
            columns = ["No", "Location", "Site Area (Ha)", "Gross Plot Ratio", "Status", "Link"]
            
            # Add headers
            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
            
            # Add data
            for row_num, site in enumerate(sites_data, 2):
                ws.cell(row=row_num, column=1).value = site.get('No', '')
                ws.cell(row=row_num, column=2).value = site.get('Location', '')
                ws.cell(row=row_num, column=3).value = site.get('Site Area (Ha)', '')
                ws.cell(row=row_num, column=4).value = site.get('Gross Plot Ratio', '')
                ws.cell(row=row_num, column=5).value = site.get('Status', '')
                ws.cell(row=row_num, column=6).value = site.get('Link', '')
            
            # Apply formatting
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the Excel file
            wb.save(self.output_file)
            logger.info(f"Successfully saved data to {self.output_file}")
            
            return True
        except Exception as e:
            logger.error(f"Error saving data: {e}")
            return False(value)
            table_df.to_excel(writer, sheet_name=sheet_name, 
                                startrow=row_position, index=False)
            row_position += len(table_df) + 3
    
    def _format_excel_sheets(self, writer):
        """Apply formatting to Excel sheets"""
        workbook = writer.book
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Apply formatting to each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Auto-adjust column widths
            for i, column in enumerate(worksheet.columns):
                max_length = 0
                column_letter = get_column_letter(i + 1)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            # Apply borders to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border
    
    def _clean_sheet_name(self, name):
        """Clean sheet name to be valid in Excel"""
        # Replace invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Truncate if too long (Excel has a 31 character limit for sheet names)
        if len(name) > 31:
            name = name[:28] + '...'
            
        return name
            
    def run(self):
        """Run the scraper"""
        logger.info("Starting URA GLS sites scraper")
        
        html = self.fetch_main_page()
        if not html:
            logger.error("Failed to fetch main page. Exiting.")
            return False
            
        sites_data = self.parse_main_page(html)
        if not sites_data:
            logger.warning("No GLS sites found. Exiting.")
            return False
            
        # Check for awarded sites and fetch their details if available
        awarded_sites = [site for site in sites_data if site['Status'].lower() == 'awarded']
        if awarded_sites:
            logger.info(f"Found {len(awarded_sites)} awarded sites. Fetching details.")
            sites_data = self.fetch_project_details(sites_data)
        else:
            logger.info("No awarded sites found.")
        
        success = self.save_data(sites_data)
        
        if success:
            logger.info("Scraping completed successfully")
        else:
            logger.error("Failed to save scraped data")
            
        return success

if __name__ == "__main__":
    scraper = URAScraper()
    scraper.run()